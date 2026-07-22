from __future__ import annotations

import argparse
import fcntl
import json
import os
import time
from pathlib import Path

import openpyxl
import pandas as pd

try:
    from quant_web import market_store
except ImportError:
    import market_store


PROJECT_ROOT = market_store.PROJECT_ROOT
MARKET_ROOT = market_store.MARKET_ROOT
DEFAULT_OUTPUT = PROJECT_ROOT / "Whole Market.xlsx"


def _project_root(root: Path) -> Path:
    return Path(root).parent.parent


def export_lock_path(root: Path = MARKET_ROOT) -> Path:
    return _project_root(Path(root)) / ".locks" / "xlsx-export.lock"


def _market_lock_path(root: Path) -> Path:
    return _project_root(root) / ".locks" / "market.lock"


def _validate_workbook(
    temp_path: Path,
    expected_sheets: list[str],
    latest_rows: int,
) -> None:
    with temp_path.open("rb") as temp_handle:
        workbook = openpyxl.load_workbook(
            temp_handle, read_only=True, data_only=True
        )
        try:
            if workbook.sheetnames != expected_sheets:
                raise ValueError(
                    "exported workbook sheet names do not match CSV snapshots"
                )
            latest_sheet = workbook[expected_sheets[-1]]
            headers = [
                cell.value
                for cell in next(latest_sheet.iter_rows(max_row=1))
            ]
            if headers != market_store.MARKET_COLUMNS:
                raise ValueError(
                    "exported workbook columns do not match market schema"
                )
            actual_rows = latest_sheet.max_row - 1
            if actual_rows != latest_rows:
                raise ValueError(
                    "exported latest row count mismatch: "
                    f"{actual_rows} != {latest_rows}"
                )
        finally:
            workbook.close()


def export_workbook(
    output_path: str | os.PathLike[str],
    root: Path = MARKET_ROOT,
) -> dict:
    started = time.monotonic()
    root = Path(root)
    output_path = Path(output_path)
    temp_path = Path(str(output_path) + ".tmp")
    lock_path = export_lock_path(root)
    lock_path.parent.mkdir(parents=True, exist_ok=True)

    with lock_path.open("a+b") as lock_handle:
        try:
            fcntl.flock(
                lock_handle.fileno(), fcntl.LOCK_EX | fcntl.LOCK_NB
            )
        except BlockingIOError:
            return {"error": "export already running"}

        try:
            with market_store.market_read_session(
                root, lock_path=_market_lock_path(root)
            ) as reader:
                full_dates = reader.list_full_dates()
                if not full_dates:
                    raise ValueError("no market snapshots to export")

                sheet_names = [full_date[4:] for full_date in full_dates]
                if len(sheet_names) != len(set(sheet_names)):
                    raise ValueError(
                        "duplicate MMDD sheet name across snapshot years"
                    )
                frames = [
                    (full_date, reader.read_snapshot(full_date))
                    for full_date in full_dates
                ]

            output_path.parent.mkdir(parents=True, exist_ok=True)
            with pd.ExcelWriter(temp_path, engine="openpyxl") as writer:
                for full_date, frame in frames:
                    frame.to_excel(
                        writer, sheet_name=full_date[4:], index=False
                    )

            latest_rows = len(frames[-1][1])
            _validate_workbook(temp_path, sheet_names, latest_rows)
            os.replace(temp_path, output_path)
            return {
                "sheets": len(frames),
                "latest": full_dates[-1],
                "rows": latest_rows,
                "elapsed": time.monotonic() - started,
            }
        finally:
            temp_path.unlink(missing_ok=True)
            fcntl.flock(lock_handle.fileno(), fcntl.LOCK_UN)


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(
        description="Rebuild the compatibility workbook from market CSV snapshots."
    )
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT)
    args = parser.parse_args(argv)
    result = export_workbook(args.output)
    print(json.dumps(result, ensure_ascii=False))
    return 1 if result.get("error") else 0


if __name__ == "__main__":
    raise SystemExit(main())
