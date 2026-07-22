import fcntl
import os
import sys
import tempfile
import unittest
from pathlib import Path
from unittest import mock

import openpyxl
import pandas as pd

sys.path.insert(0, str(Path(__file__).resolve().parent))

from quant_web import app as market_app
from quant_web import export_market_xlsx
from quant_web import market_store


def sample_market(rows):
    return pd.DataFrame({
        "全称": [f"股票{i}" for i in range(rows)],
        "代码": [f"{i:06d}.SZ" for i in range(rows)],
        "行业": ["银行"] * rows,
        "地域": ["深圳"] * rows,
        "准确率": [0.7] * rows,
        "追踪天数": [90] * rows,
        "倍数": ["1.48(5.05)"] * rows,
        "指标历史": ["(+1.20%)"] * rows,
        "近三日涨幅": ["-"] * rows,
        "今日指标": [1.2] * rows,
        "指标趋势": [0.2] * rows,
    })


class TestMarketWorkbookExport(unittest.TestCase):
    def setUp(self):
        self.temp_dir = tempfile.TemporaryDirectory()
        self.project_root = Path(self.temp_dir.name)
        self.market_root = self.project_root / "data" / "market"
        self.market_lock = self.project_root / ".locks" / "market.lock"
        self.output = self.project_root / "Whole Market.xlsx"

    def tearDown(self):
        self.temp_dir.cleanup()

    def write_snapshot(self, full_date, frame):
        market_store.write_snapshot(
            full_date,
            frame,
            root=self.market_root,
            lock_path=self.market_lock,
            min_rows=1,
        )

    def test_export_preserves_order_columns_rows_and_csv_parity(self):
        frames = {
            "20260331": sample_market(2),
            "20260401": sample_market(3),
        }
        for full_date, frame in reversed(list(frames.items())):
            self.write_snapshot(full_date, frame)

        result = export_market_xlsx.export_workbook(
            self.output, root=self.market_root
        )

        workbook = openpyxl.load_workbook(
            self.output, read_only=True, data_only=True
        )
        try:
            self.assertEqual(workbook.sheetnames, ["0331", "0401"])
            self.assertEqual(workbook["0331"].max_column, 11)
            self.assertEqual(workbook["0401"].max_column, 11)
            self.assertEqual(workbook["0331"].max_row, 3)
            self.assertEqual(workbook["0401"].max_row, 4)
        finally:
            workbook.close()

        for full_date, expected in frames.items():
            actual = pd.read_excel(self.output, sheet_name=full_date[4:])
            market_store.compare_frames(expected, actual)
        self.assertEqual(result["sheets"], 2)
        self.assertEqual(result["latest"], "20260401")
        self.assertEqual(result["rows"], 3)
        self.assertGreaterEqual(result["elapsed"], 0)
        self.assertFalse(Path(str(self.output) + ".tmp").exists())

    def test_export_refuses_duplicate_mmdd_across_years(self):
        self.write_snapshot("20260331", sample_market(2))
        corrupt_path = self.market_root / "2027" / "20270331.csv"
        corrupt_path.parent.mkdir(parents=True)
        sample_market(2).to_csv(
            corrupt_path, index=False, encoding="utf-8-sig"
        )

        with self.assertRaisesRegex(ValueError, "duplicate MMDD sheet name"):
            export_market_xlsx.export_workbook(
                self.output, root=self.market_root
            )

        self.assertFalse(self.output.exists())
        self.assertFalse(Path(str(self.output) + ".tmp").exists())

    def test_export_uses_one_shared_market_read_session(self):
        self.write_snapshot("20260331", sample_market(2))
        real_session = market_store.market_read_session
        with mock.patch.object(
            export_market_xlsx.market_store,
            "market_read_session",
            wraps=real_session,
        ) as read_session:
            export_market_xlsx.export_workbook(
                self.output, root=self.market_root
            )

        self.assertEqual(read_session.call_count, 1)

    def test_busy_export_lock_returns_clear_result_without_blocking(self):
        self.write_snapshot("20260331", sample_market(2))
        export_lock = export_market_xlsx.export_lock_path(self.market_root)
        self.assertEqual(
            export_lock,
            self.project_root / ".locks" / "xlsx-export.lock",
        )
        export_lock.parent.mkdir(parents=True, exist_ok=True)
        with export_lock.open("a+b") as handle:
            fcntl.flock(handle.fileno(), fcntl.LOCK_EX | fcntl.LOCK_NB)
            result = export_market_xlsx.export_workbook(
                self.output, root=self.market_root
            )

        self.assertEqual(result, {"error": "export already running"})
        self.assertFalse(self.output.exists())

    def test_failed_validation_preserves_output_and_removes_only_temp(self):
        self.write_snapshot("20260331", sample_market(2))
        self.output.write_bytes(b"existing workbook")
        unrelated = self.project_root / "keep.tmp"
        unrelated.write_text("keep", encoding="utf-8")

        with mock.patch.object(
            export_market_xlsx.openpyxl,
            "load_workbook",
            side_effect=ValueError("validation failed"),
        ):
            with self.assertRaisesRegex(ValueError, "validation failed"):
                export_market_xlsx.export_workbook(
                    self.output, root=self.market_root
                )

        self.assertEqual(self.output.read_bytes(), b"existing workbook")
        self.assertFalse(Path(str(self.output) + ".tmp").exists())
        self.assertEqual(unrelated.read_text(encoding="utf-8"), "keep")


class TestMarketExportAdminAction(unittest.TestCase):
    def invoke(self, action):
        with market_app.app.test_request_context(
            "/admin/health/action/" + action
        ):
            return market_app.admin_health_action.__wrapped__(action)

    def test_admin_action_launches_detached_exporter_and_reports_start(self):
        with mock.patch.object(market_app.subprocess, "Popen") as popen, \
                mock.patch.object(market_app, "flash") as flash:
            response = self.invoke("export_market_xlsx")

        popen.assert_called_once_with(
            [
                sys.executable,
                os.path.join(
                    market_app._PROJECT_ROOT,
                    "quant_web",
                    "export_market_xlsx.py",
                ),
                "--output",
                os.path.join(market_app._PROJECT_ROOT, "Whole Market.xlsx"),
            ],
            start_new_session=True,
            stdout=market_app.subprocess.DEVNULL,
            stderr=market_app.subprocess.DEVNULL,
        )
        flash.assert_called_once_with("Workbook generation started")
        self.assertEqual(response.status_code, 302)

    def test_health_template_shows_csv_and_export_stats(self):
        template = (
            Path(market_app.__file__).parent / "templates" / "admin_health.html"
        ).read_text(encoding="utf-8")

        for field in (
            "csv_files",
            "csv_size_mb",
            "latest_csv",
            "xlsx_generated_at",
        ):
            self.assertIn(f"data_stats.{field}", template)


if __name__ == "__main__":
    unittest.main()
