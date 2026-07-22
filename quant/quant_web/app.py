"""Flask 路由层 — 数据由 data_service 提供。"""
import functools
import hashlib
import hmac
import io
import json
import os
import random
import re
import secrets
import smtplib
import sqlite3
import sys
import threading
import time
import urllib.request
from datetime import datetime
import glob
import subprocess
import pandas as pd
from email.mime.text import MIMEText

from flask import Flask, render_template, jsonify, request, Response, session, redirect, url_for, flash
from werkzeug.middleware.proxy_fix import ProxyFix

_PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
_PIPELINE_STATUS_PATH = os.path.join(_PROJECT_ROOT, "pipeline_status.json")

import data_service as ds
import db_service as dbs
import market_store
import market_writer
import refresh_market_cache as market_cache_refresher
from report_paths import glob_reports, resolve_report

app = Flask(__name__)
app.secret_key = os.environ.get("FLASK_SECRET_KEY", "quant-web-secret-change-in-production")
app.url_map.strict_slashes = False

# 告知 Flask 它在 nginx 反向代理后面，信任 X-Forwarded-* 头
app.wsgi_app = ProxyFix(app.wsgi_app, x_for=1, x_proto=1)

app.config.update(
    SESSION_COOKIE_SAMESITE="Lax",
    SESSION_COOKIE_SECURE=True,
    SESSION_COOKIE_HTTPONLY=True,
)

# ── Jinja 自定义过滤器 ──
@app.template_filter("pct")
def _pct_filter(v):
    """将小数格式化为百分比字符串，如 0.0996 → '9.96%'"""
    try:
        f = float(v)
        if f == int(f):
            return f"{int(f):d}%"
        return f"{f * 100:.2f}%"
    except (ValueError, TypeError):
        return v

_LOGIN_LOG_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "login_log.json")

# ── SMTP 配置 ──
_SMTP_USER = os.environ.get("SMTP_USER", "")
_SMTP_PASS = os.environ.get("SMTP_PASS", "")
_SMTP_HOST = os.environ.get("SMTP_HOST", "")
_SMTP_PORT = int(os.environ.get("SMTP_PORT", "465"))

# ── 邀请码 ──
_INVITE_CODE = os.environ.get("INVITE_CODE", "CHANGE_ME")

# ── 用户数据库（SQLite） ──
_DB_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "users.db")

# 管理员/用户初始密码（从环境变量读取，部署时设置）
_ADMIN_PASSWORD = os.environ.get("ADMIN_PASSWORD", "")
_USER_PASSWORD = os.environ.get("USER_PASSWORD", "")


def _hash_password(password: str) -> str:
    """pbkdf2_hmac 哈希密码。"""
    salt = os.urandom(16)
    dk = hashlib.pbkdf2_hmac("sha256", password.encode(), salt, 100000)
    return salt.hex() + ":" + dk.hex()


def _verify_password(password: str, stored: str) -> bool:
    """验证密码是否匹配哈希值。"""
    try:
        salt_hex, dk_hex = stored.split(":")
        salt = bytes.fromhex(salt_hex)
        dk = hashlib.pbkdf2_hmac("sha256", password.encode(), salt, 100000)
        return dk.hex() == dk_hex
    except Exception:
        return False


def _db_get():
    """获取 SQLite 连接（同一线程复用）。"""
    conn = sqlite3.connect(_DB_PATH)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA journal_mode=WAL")
    return conn


def _db_init():
    """建表 + 迁移旧用户（幂等）。"""
    conn = _db_get()
    conn.execute("""CREATE TABLE IF NOT EXISTS users (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        username TEXT UNIQUE NOT NULL,
        email TEXT UNIQUE NOT NULL DEFAULT '',
        password_hash TEXT NOT NULL,
        role TEXT DEFAULT 'user',
        name TEXT DEFAULT '',
        created_at TEXT DEFAULT (datetime('now', 'localtime'))
    )""")
    conn.commit()

    # 创建初始用户（密码从环境变量读取，已存在则跳过）
    defaults = [
        ("admin", _ADMIN_PASSWORD, "admin", "超级管理员"),
        ("user", _USER_PASSWORD, "user", "普通用户"),
    ]
    for uname, pw, role, name in defaults:
        if not pw:
            continue
        cur = conn.execute("SELECT id FROM users WHERE username = ?", (uname,))
        if cur.fetchone() is None:
            h = _hash_password(pw)
            conn.execute(
                "INSERT INTO users (username, email, password_hash, role, name) VALUES (?, ?, ?, ?, ?)",
                (uname, uname + "@local", h, role, name))
    conn.commit()
    conn.close()


def _find_user(username: str) -> dict | None:
    """根据用户名查找用户，返回 dict 或 None。"""
    conn = _db_get()
    row = conn.execute(
        "SELECT username, email, password_hash, role, name FROM users WHERE username = ?",
        (username,)
    ).fetchone()
    conn.close()
    if row is None:
        return None
    return {
        "username": row["username"],
        "email": row["email"],
        "password_hash": row["password_hash"],
        "role": row["role"],
        "name": row["name"] or row["username"],
    }


def _find_user_by_email_prefix(prefix: str) -> dict | None:
    """根据邮箱前缀查找用户，仅当唯一匹配时返回，多条匹配视为歧义。"""
    conn = _db_get()
    rows = conn.execute(
        "SELECT username, email, password_hash, role, name FROM users WHERE email LIKE ? LIMIT 2",
        (prefix + '@%',)
    ).fetchall()
    conn.close()
    if len(rows) != 1:
        return None  # 0 条未找到，2 条前缀歧义
    row = rows[0]
    return {
        "username": row["username"],
        "email": row["email"],
        "password_hash": row["password_hash"],
        "role": row["role"],
        "name": row["name"] or row["username"],
    }


def _find_user_by_email(email: str) -> dict | None:
    """根据邮箱查找用户。"""
    conn = _db_get()
    row = conn.execute(
        "SELECT username, email, password_hash, role, name FROM users WHERE email = ?",
        (email,)
    ).fetchone()
    conn.close()
    if row is None:
        return None
    return {
        "username": row["username"],
        "email": row["email"],
        "password_hash": row["password_hash"],
        "role": row["role"],
        "name": row["name"] or row["username"],
    }


def _create_user(username: str, email: str, password: str,
                 role: str = "user", name: str = "") -> tuple[bool, str]:
    """创建新用户。返回 (成功, 错误信息)。"""
    conn = _db_get()
    try:
        conn.execute(
            "INSERT INTO users (username, email, password_hash, role, name) VALUES (?, ?, ?, ?, ?)",
            (username, email, _hash_password(password), role, name or username)
        )
        conn.commit()
        return True, ""
    except sqlite3.IntegrityError as e:
        msg = str(e).lower()
        if "username" in msg:
            return False, "用户名已被注册"
        elif "email" in msg:
            return False, "该邮箱已被注册"
        return False, "注册失败，请重试"
    finally:
        conn.close()


def _list_users() -> list[dict]:
    """列出所有用户（按创建时间倒序）。"""
    conn = _db_get()
    rows = conn.execute(
        "SELECT id, username, email, role, name, created_at FROM users ORDER BY created_at DESC"
    ).fetchall()
    conn.close()
    return [{"id": r["id"], "username": r["username"], "email": r["email"],
             "role": r["role"], "name": r["name"], "created_at": r["created_at"]} for r in rows]


def _update_user_role(username: str, role: str) -> bool:
    """修改用户角色。"""
    conn = _db_get()
    conn.execute("UPDATE users SET role = ? WHERE username = ?", (role, username))
    conn.commit()
    affected = conn.total_changes
    conn.close()
    return affected > 0


def _delete_user(username: str) -> bool:
    """删除用户（不可删除自己）。"""
    conn = _db_get()
    conn.execute("DELETE FROM users WHERE username = ?", (username,))
    conn.commit()
    affected = conn.total_changes
    conn.close()
    return affected > 0


def _reset_user_password(username: str, new_password: str) -> bool:
    """重置用户密码。"""
    conn = _db_get()
    h = _hash_password(new_password)
    conn.execute("UPDATE users SET password_hash = ? WHERE username = ?", (h, username))
    conn.commit()
    affected = conn.total_changes
    conn.close()
    return affected > 0


# ── 验证码（SQLite 存储，多 worker 共享） ──

def _init_code_table(conn):
    conn.execute("""CREATE TABLE IF NOT EXISTS verify_codes (
        email TEXT PRIMARY KEY,
        code TEXT NOT NULL,
        expires REAL NOT NULL,
        attempts INTEGER DEFAULT 0,
        last_sent REAL NOT NULL
    )""")
    conn.commit()


def _gen_code() -> str:
    return str(random.randint(100000, 999999))


def _set_code(email: str, code: str):
    conn = _db_get()
    _init_code_table(conn)
    now = time.time()
    conn.execute(
        "INSERT OR REPLACE INTO verify_codes (email, code, expires, attempts, last_sent) VALUES (?, ?, ?, 0, ?)",
        (email, code, now + 300, now)
    )
    conn.commit()
    conn.close()


def _check_code(email: str, code: str) -> str:
    """验证验证码。返回空字符串表示正确，否则返回错误信息。"""
    conn = _db_get()
    _init_code_table(conn)
    now = time.time()
    # 清理过期
    conn.execute("DELETE FROM verify_codes WHERE expires < ?", (now,))
    conn.commit()

    row = conn.execute(
        "SELECT code, expires, attempts FROM verify_codes WHERE email = ?", (email,)
    ).fetchone()
    if row is None:
        conn.close()
        return "请先获取验证码"
    if row["expires"] < now:
        conn.execute("DELETE FROM verify_codes WHERE email = ?", (email,))
        conn.commit()
        conn.close()
        return "验证码已过期，请重新获取"
    if row["attempts"] >= 5:
        conn.execute("DELETE FROM verify_codes WHERE email = ?", (email,))
        conn.commit()
        conn.close()
        return "验证码尝试次数过多，请重新获取"
    conn.execute("UPDATE verify_codes SET attempts = attempts + 1 WHERE email = ?", (email,))
    conn.commit()
    if row["code"] != code:
        conn.close()
        return "验证码错误"
    conn.execute("DELETE FROM verify_codes WHERE email = ?", (email,))
    conn.commit()
    conn.close()
    return ""


def _can_resend(email: str) -> str:
    """检查是否可重发。返回空字符串表示可以，否则返回错误信息。"""
    conn = _db_get()
    _init_code_table(conn)
    now = time.time()
    row = conn.execute(
        "SELECT last_sent FROM verify_codes WHERE email = ?", (email,)
    ).fetchone()
    conn.close()
    if row and now - row["last_sent"] < 60:
        remain = 60 - int(now - row["last_sent"])
        return f"请 {remain} 秒后再试"
    return ""


def _send_code(email_to: str, code: str) -> tuple[bool, str]:
    """通过 QQ SMTP 发送验证码。返回 (成功, 错误信息)。"""
    try:
        msg = MIMEText(
            f"您的量化分析系统注册验证码是：{code}，请在 5 分钟内完成验证。\n\n"
            f"如非本人操作，请忽略此邮件。",
            "plain", "utf-8"
        )
        msg["Subject"] = "量化分析系统 — 注册验证码"
        msg["From"] = _SMTP_USER
        msg["To"] = email_to

        with smtplib.SMTP_SSL(_SMTP_HOST, _SMTP_PORT, timeout=10) as s:
            s.login(_SMTP_USER, _SMTP_PASS)
            s.send_message(msg)
        return True, ""
    except smtplib.SMTPAuthenticationError:
        return False, "邮件发送失败：SMTP 认证错误，请联系管理员检查邮箱配置"
    except smtplib.SMTPException as e:
        return False, f"邮件发送失败：{e}"
    except Exception as e:
        return False, f"邮件发送异常：{e}"


# 应用启动时初始化数据库
_db_init()


def _read_pipeline_status():
    """Read pipeline_status.json, return dict with defaults if missing."""
    try:
        with open(_PIPELINE_STATUS_PATH, 'r', encoding='utf-8') as f:
            return json.load(f)
    except (FileNotFoundError, json.JSONDecodeError):
        return {"running": False, "history": []}


def _write_pipeline_status(data):
    """Write pipeline_status.json atomically."""
    tmp = _PIPELINE_STATUS_PATH + ".tmp"
    with open(tmp, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    os.replace(tmp, _PIPELINE_STATUS_PATH)


def _write_sheet_to_master(master_path, df, sheet_name):
    """Write a DataFrame as a sheet to Whole Market.xlsx with auto-repair on error."""
    import shutil
    if not os.path.exists(master_path):
        df.to_excel(master_path, sheet_name=sheet_name, index=False)
        return
    try:
        _openpyxl_append_sheet(master_path, df, sheet_name)
    except Exception:
        # Auto-repair ZIP, then retry
        bak = master_path + '.auto_repair.bak'
        shutil.copy2(master_path, bak)
        try:
            _repair_xlsx_zip(master_path)
            _openpyxl_append_sheet(master_path, df, sheet_name)
        except Exception:
            shutil.copy2(bak, master_path)
            raise


def _refresh_market_cache(full_date):
    """Best-effort refresh of market.db from the configured market storage."""
    try:
        rows = market_cache_refresher.refresh_market_cache(full_date)
        print(f"[pipeline] SQLite market cache refreshed: {full_date} ({rows} rows)")
        return True
    except Exception as e:
        print(f"[pipeline] SQLite market cache refresh failed: {e}")
        return False


def _openpyxl_append_sheet(path, df, sheet_name):
    """Append/replace a sheet using openpyxl with atomic write (safe on crash)."""
    from openpyxl import load_workbook
    from openpyxl.utils.dataframe import dataframe_to_rows
    wb = load_workbook(path, keep_vba=False, keep_links=False)
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name)
    for row in dataframe_to_rows(df, index=False, header=True):
        ws.append(row)
    tmp = path + '.tmp'
    wb.save(tmp)
    wb.close()
    os.replace(tmp, path)  # atomic — interrupt here won't corrupt the original


def _repair_xlsx_zip(path):
    """Fix broken xlsx by scanning ZIP local file headers (works even without central directory)."""
    import struct, zlib
    with open(path, 'rb') as f:
        data = f.read()

    entries = []
    pos = 0
    while pos < len(data) - 30:
        if data[pos:pos+4] != b'PK\x03\x04':
            pos += 1
            continue
        compression = struct.unpack('<H', data[pos+8:pos+10])[0]
        comp_size = struct.unpack('<I', data[pos+18:pos+22])[0]
        name_len = struct.unpack('<H', data[pos+26:pos+28])[0]
        extra_len = struct.unpack('<H', data[pos+28:pos+30])[0]
        name = data[pos+30:pos+30+name_len].decode('latin-1')
        start = pos + 30 + name_len + extra_len
        raw = data[start:start+comp_size]
        if compression == 8:
            try:
                content = zlib.decompress(raw, -zlib.MAX_WBITS)
            except Exception:
                content = _minimal_ooxml_part(name)
        elif compression == 0:
            content = raw
        else:
            content = _minimal_ooxml_part(name)
        entries.append((name, content))
        pos = start + comp_size

    if not entries:
        raise ValueError("No ZIP entries found, cannot repair")

    import io as _io, zipfile
    fixed = _io.BytesIO()
    with zipfile.ZipFile(fixed, 'w', zipfile.ZIP_DEFLATED) as zout:
        for name, content in entries:
            zout.writestr(name, content)

    with open(path, 'wb') as f:
        f.write(fixed.getvalue())


def _minimal_ooxml_part(filename):
    """Return minimal valid XML content for common OOXML parts."""
    core = b'''<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<cp:coreProperties xmlns:cp="http://schemas.openxmlformats.org/package/2006/metadata/core-properties">
<dc:creator xmlns:dc="http://purl.org/dc/elements/1.1/">quant</dc:creator>
</cp:coreProperties>'''
    app = b'''<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Properties xmlns="http://schemas.openxmlformats.org/officeDocument/2006/extended-properties"/>'''
    if 'core.xml' in filename:
        return core
    if 'app.xml' in filename:
        return app
    return b'<?xml version="1.0"?><root/>'


def _build_pipeline_email(date_str, backtest_ok, weekly_ok, elapsed):
    """Build email body with backtest + industry + weekly settlement results."""
    lines = []
    lines.append(f"量化流水线 {date_str[:2]}/{date_str[2:]} 执行报告")
    status = "成功" if (backtest_ok and weekly_ok) else "失败"
    lines.append(f"状态: {status}")
    lines.append("")

    # ── Backtest results ──
    bt_json = resolve_report(f"{date_str}量化复盘报告.json")
    if os.path.exists(bt_json):
        try:
            with open(bt_json, 'r', encoding='utf-8') as f:
                bt = json.load(f)
            t = bt.get('tracking', {})
            lines.append("━━ 回测 ━━")
            lines.append(f"全样本胜率: {t.get('all_raw', 'N/A')}")
            lines.append(f"冷门Alpha: {t.get('cold_alpha_raw', 'N/A')}")
            lines.append(f"明日>1.0 达标: {t.get('next_10', 'N/A')} 只")

            sheets = bt.get('sheets', {})
            ind_dist = sheets.get('industry_dist', {})
            ind_rows = ind_dist.get('rows', [])
            if ind_rows:
                lines.append("")
                prev_date_str = f"{date_str[:2]}/{date_str[2:]}"
                lines.append(f"昨日行业明细 ({prev_date_str}):")
                for row in ind_rows:
                    name = row.get('行业名称', '?')
                    count = row.get('入选数量', 0)
                    ret = row.get('今日收益率', 0)
                    stocks = row.get('包含股票名单', '')
                    try:
                        ret_f = float(ret) * 100
                        ret_str = f"{ret_f:+.2f}%"
                    except (ValueError, TypeError):
                        ret_str = str(ret)
                    lines.append(f"  {name}  {count}只  平均收益 {ret_str}  股票: {stocks}")
        except Exception as e:
            lines.append(f"(回测结果读取失败: {e})")
    else:
        lines.append("━━ 回测: 失败 ━━")

    lines.append("")

    # ── Five-days-ago weekly settlement ──
    try:
        all_sheets = sorted(dbs.get_market_dates(), key=lambda s: int(s))
        bt_dates_set = set()
        for f in glob_reports("*量化复盘报告.xlsx"):
            d = os.path.basename(f)[:4]
            if d.isdigit():
                bt_dates_set.add(d)
        ordered_dates = [d for d in all_sheets if d in bt_dates_set]
        if date_str in ordered_dates:
            idx = ordered_dates.index(date_str)
            settle_idx = idx - 5
            if settle_idx >= 0:
                settle_date = ordered_dates[settle_idx]
                wk_json = resolve_report(f"{settle_date}的选股策略礼拜攻势.json")
                if os.path.exists(wk_json):
                    with open(wk_json, 'r', encoding='utf-8') as f:
                        wk = json.load(f)
                    lines.append(f"━━ 礼拜攻势 {settle_date[:2]}/{settle_date[2:]} 结仓(5日) ━━")
                    std_summary = wk.get('standard', {}).get('summary', [])
                    for s in std_summary:
                        lines.append(f"  {s.get('策略分组', '?')}: {s.get('入选股票数', 0)}只  平均回报 {s.get('平均持仓回报', '?')}  胜率 {s.get('策略胜率(>0%)', '?')}")
                    top_summary = wk.get('top_industries', {}).get('summary', [])
                    if top_summary:
                        top_strs = []
                        for ts in top_summary:
                            ret = ts.get('平均持仓回报', 0)
                            try:
                                ret_f = float(ret) * 100
                                ret_s = f"{ret_f:+.1f}%"
                            except (ValueError, TypeError):
                                ret_s = str(ret)
                            top_strs.append(f"{ts.get('策略分组', '?')}({ret_s})")
                        lines.append(f"  前三行业: {', '.join(top_strs)}")
                    cold_summary = wk.get('cold_industry', {}).get('summary', [])
                    if cold_summary:
                        cold_strs = []
                        for cs in cold_summary:
                            ret = cs.get('平均持仓回报', 0)
                            try:
                                ret_f = float(ret) * 100
                                ret_s = f"{ret_f:+.1f}%"
                            except (ValueError, TypeError):
                                ret_s = str(ret)
                            cold_strs.append(f"{cs.get('策略分组', '?')}({ret_s})")
                        lines.append(f"  冷门行业: {', '.join(cold_strs)}")
                else:
                    lines.append(f"━━ 礼拜攻势 {settle_date[:2]}/{settle_date[2:]} 结仓: 无数据 ━━")
            else:
                lines.append("━━ 五日前无结仓 ━━")
    except Exception:
        pass

    lines.append("")

    # ── New weekly generation ──
    wk_new_json = resolve_report(f"{date_str}的选股策略礼拜攻势.json")
    if os.path.exists(wk_new_json):
        try:
            with open(wk_new_json, 'r', encoding='utf-8') as f:
                wk_new = json.load(f)
            n_avail = wk_new.get('meta', {}).get('n_available', 0)
            std_new = wk_new.get('standard', {}).get('summary', [])
            lines.append(f"━━ 礼拜攻势 {date_str[:2]}/{date_str[2:]} 已生成 ({n_avail}/5 天可用) ━━")
            for s in std_new:
                lines.append(f"  {s.get('策略分组', '?')}: {s.get('入选股票数', 0)}只")
        except Exception:
            pass

    lines.append("")
    lines.append(f"总耗时: {elapsed:.0f} 秒")
    lines.append("")
    lines.append("—— 量化分析系统自动通知")

    return "\n".join(lines)


def _run_pipeline_for_date(date_yyyymmdd, temp_csv=None):
    """Background thread: persist market data, then run backtest and weekly jobs."""
    import subprocess
    import sys
    import pandas as pd
    from datetime import datetime

    date_mmdd = date_yyyymmdd[4:]  # "20260511" -> "0511"

    status = _read_pipeline_status()
    status["running"] = True
    status["date"] = date_mmdd
    status["stage"] = "write_master"
    status["started_at"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    _write_pipeline_status(status)

    start = datetime.now()

    try:
        # Step 0: 写入 Master（从临时 CSV 读取，避免上传接口等待）
        if temp_csv and os.path.exists(temp_csv):
            df = pd.read_csv(temp_csv)
            market_writer.persist_market_dataframe(date_yyyymmdd, df)
            status["stage"] = "refresh_market_cache"
            _write_pipeline_status(status)
            if not _refresh_market_cache(date_yyyymmdd):
                raise RuntimeError("SQLite market cache refresh failed; reports not started")
            os.remove(temp_csv)

        # Run backtest
        status["stage"] = "backtest"
        _write_pipeline_status(status)
        bt_script = os.path.join(os.path.dirname(os.path.abspath(__file__)), "batch_backtest.py")
        bt_result = subprocess.run(
            [sys.executable, bt_script, "--date", date_mmdd],
            capture_output=True, text=True, timeout=300,
            cwd=os.path.dirname(os.path.abspath(__file__)),
        )
        backtest_ok = bt_result.returncode == 0
        status["backtest_returncode"] = bt_result.returncode
        if bt_result.returncode != 0:
            status["backtest_error"] = (bt_result.stderr or bt_result.stdout or "")[-4000:]

        # Run weekly
        status["stage"] = "weekly"
        _write_pipeline_status(status)

        wk_script = os.path.join(os.path.dirname(os.path.abspath(__file__)), "batch_weekly.py")
        wk_result = subprocess.run(
            [sys.executable, wk_script, "--date", date_mmdd],
            capture_output=True, text=True, timeout=300,
            cwd=os.path.dirname(os.path.abspath(__file__)),
        )
        weekly_ok = wk_result.returncode == 0
        status["weekly_returncode"] = wk_result.returncode
        if wk_result.returncode != 0:
            status["weekly_error"] = (wk_result.stderr or wk_result.stdout or "")[-4000:]

        elapsed = (datetime.now() - start).total_seconds()

        # Update status
        status["running"] = False
        status["stage"] = "done"
        status["finished_at"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        status["backtest_ok"] = backtest_ok
        status["weekly_ok"] = weekly_ok
        status.setdefault("history", []).insert(0, {
            "date": date_mmdd,
            "status": "success" if (backtest_ok and weekly_ok) else "fail",
            "time": status["started_at"],
            "elapsed": int(elapsed),
        })
        status["history"] = status["history"][:20]
        _write_pipeline_status(status)

        # Send email
        body = _build_pipeline_email(date_mmdd, backtest_ok, weekly_ok, elapsed)
        _send_pipeline_notification(backtest_ok and weekly_ok, body)
    except Exception as e:
        elapsed = (datetime.now() - start).total_seconds()
        if temp_csv and os.path.exists(temp_csv):
            try:
                os.remove(temp_csv)
            except OSError:
                pass
        status["running"] = False
        status["stage"] = "failed"
        status["finished_at"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        status["error"] = str(e)
        status.setdefault("history", []).insert(0, {
            "date": date_mmdd,
            "status": "fail",
            "time": status.get("started_at"),
            "elapsed": int(elapsed),
        })
        status["history"] = status["history"][:20]
        _write_pipeline_status(status)
        _send_pipeline_notification(False, f"上传流水线 {date_mmdd} 失败: {e}")


def _send_pipeline_notification(success, body):
    """Send pipeline completion email, or log if SMTP unavailable."""
    import logging
    if not (_SMTP_USER and _SMTP_PASS and _SMTP_HOST):
        logging.getLogger(__name__).info("SMTP配置不完整，跳过流水线通知")
        return
    try:
        date_str = datetime.now().strftime("%m/%d")
        status_label = "成功" if success else "失败"
        msg = MIMEText(body, "plain", "utf-8")
        msg["Subject"] = f"量化流水线 {date_str} {status_label}"
        msg["From"] = _SMTP_USER
        notify_to = os.environ.get("NOTIFY_EMAIL") or _SMTP_USER
        msg["To"] = notify_to
        with smtplib.SMTP_SSL(_SMTP_HOST, _SMTP_PORT, timeout=10) as s:
            s.login(_SMTP_USER, _SMTP_PASS)
            s.send_message(msg)
    except Exception:
        logging.getLogger(__name__).warning("流水线邮件发送失败", exc_info=True)


def _parse_ua(ua: str) -> dict:
    """从 User-Agent 解析浏览器、操作系统和设备信息。"""
    info = {"browser": "", "browser_ver": "", "os": "", "os_ver": "", "device": "", "model": ""}
    if not ua:
        return info

    ua_lower = ua.lower()

    # ── 浏览器 ──
    for key, name in [("edg/", "Edge"), ("opr/", "Opera"), ("chrome/", "Chrome"),
                       ("firefox/", "Firefox"), ("safari/", "Safari")]:
        if key in ua_lower:
            info["browser"] = name
            m = re.search(rf'{key}(\S+)', ua_lower)
            if m:
                info["browser_ver"] = m.group(1).rstrip('.')
            break

    # ── 操作系统 ──
    os_map = [
        (r'windows nt (\d+\.?\d*)', 'Windows'),
        (r'mac os x (\S+)', 'macOS'),
        (r'android (\S+)', 'Android'),
        (r'like mac os x', 'iOS'),
        (r'linux', 'Linux'),
    ]
    for pat, name in os_map:
        m = re.search(pat, ua_lower)
        if m:
            info["os"] = name
            info["os_ver"] = m.group(1) if m.lastindex and m.lastindex >= 1 else ""
            # Windows NT version → readable name
            if name == "Windows" and info["os_ver"]:
                ver_map = {"10.0": "10/11", "6.3": "8.1", "6.2": "8", "6.1": "7"}
                info["os_ver"] = ver_map.get(info["os_ver"], info["os_ver"])
            break

    # ── 设备类型与型号 ──
    if re.search(r'iphone', ua_lower):
        info["device"] = "手机"
        m = re.search(r'iphone[\d\s,]+', ua)
        info["model"] = m.group(0).strip() if m else "iPhone"
    elif re.search(r'ipad', ua_lower):
        info["device"] = "平板"
        info["model"] = "iPad"
    elif re.search(r'android', ua_lower):
        m = re.search(r';(.+?)\)', ua)
        raw = m.group(1).strip() if m else ""
        parts = [p.strip() for p in raw.split(';') if p.strip()]
        info["model"] = parts[-1] if parts else ""
        info["device"] = "手机" if re.search(r'mobile', ua_lower) else "平板"
    elif re.search(r'windows|mac os|linux', ua_lower):
        info["device"] = "桌面电脑"
        if re.search(r'windows', ua_lower):
            info["model"] = "Windows PC"
        elif re.search(r'macintosh|mac os', ua_lower):
            info["model"] = "Mac"
        else:
            info["model"] = "Linux"

    if not info["device"]:
        info["device"] = "桌面电脑" if re.search(r'windows|macintosh|linux', ua_lower) else "未知"

    return info


def _build_device_name(ua_info: dict | None, client_info: dict | None) -> str:
    """组合设备名：操作系统 + 设备类型。"""
    ua = ua_info or {}
    os = ua.get("os", "")
    os_ver = ua.get("os_ver", "")
    device = ua.get("device", "")
    parts = [p for p in [os, os_ver, device] if p]
    return " ".join(parts) if parts else "未知设备"


def _build_device_model(ua_info: dict | None, client_info: dict | None) -> str:
    """组合设备型号：优先用 userAgentData 高熵值，其次 UA 解析结果。"""
    ci = client_info or {}
    ua = ua_info or {}

    # Chrome 90+ userAgentData 高熵值（含设备型号）
    if ci.get("brands"):
        try:
            brands_data = json.loads(ci["brands"])
            m = brands_data.get("model", "")
            if m:
                return m
        except Exception:
            pass

    # UA 解析的 model
    ua_model = ua.get("model", "")
    if ua_model:
        return ua_model

    # 桌面端回退到 navigator.platform
    platform = ci.get("platform", "")
    return platform or "未知"


def _geo_lookup(ip: str) -> str:
    """通过 ip-api.com 获取 IP 地理位置（超时 3 秒，失败返回空）。"""
    if not ip:
        return ""
    # 内网 IP 不查询
    if ip.startswith(("127.", "10.", "172.16.", "172.17.", "172.18.", "172.19.",
                      "172.20.", "172.21.", "172.22.", "172.23.", "172.24.",
                      "172.25.", "172.26.", "172.27.", "172.28.", "172.29.",
                      "172.30.", "172.31.", "192.168.", "::1")):
        return "内网 / 本地"
    try:
        url = f"http://ip-api.com/json/{ip}?lang=zh-CN&fields=status,country,city,isp,query"
        with urllib.request.urlopen(url, timeout=3) as resp:
            data = json.loads(resp.read().decode())
        if data.get("status") == "success":
            parts = [data.get("country", ""), data.get("city", "")]
            isp = data.get("isp", "")
            loc = " ".join(p for p in parts if p)
            if isp:
                loc += f" | {isp}"
            return loc
    except Exception:
        pass
    return ""


def _log_login(username: str, name: str, ip: str,
               ua_info: dict | None = None, client_info: dict | None = None) -> None:
    """记录登录事件到 JSON 文件，含设备/位置等详细信息。"""
    records = []
    if os.path.exists(_LOGIN_LOG_PATH):
        try:
            with open(_LOGIN_LOG_PATH, "r", encoding="utf-8") as f:
                records = json.load(f)
        except Exception:
            records = []

    entry = {
        "username": username,
        "name": name,
        "ip": ip,
        "time": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "location": _geo_lookup(ip),
        "browser": (ua_info or {}).get("browser", ""),
        "browser_ver": (ua_info or {}).get("browser_ver", ""),
        "os": (ua_info or {}).get("os", ""),
        "os_ver": (ua_info or {}).get("os_ver", ""),
        "device": (ua_info or {}).get("device", ""),
        "device_name": _build_device_name(ua_info, client_info),
        "model": _build_device_model(ua_info, client_info),
        "screen": (client_info or {}).get("screen", ""),
        "cores": (client_info or {}).get("cores", ""),
        "memory": (client_info or {}).get("memory", ""),
        "timezone": (client_info or {}).get("timezone", ""),
        "platform": (client_info or {}).get("platform", ""),
        "language": (client_info or {}).get("language", ""),
    }
    records.append(entry)
    records = records[-500:]
    try:
        with open(_LOGIN_LOG_PATH, "w", encoding="utf-8") as f:
            json.dump(records, f, ensure_ascii=False, indent=2)
    except Exception:
        pass


def login_required(f):
    """页面路由保护：未登录跳转到登录页。"""
    @functools.wraps(f)
    def wrapper(*args, **kwargs):
        if "user" not in session:
            if request.is_json or request.path.startswith("/api/"):
                return jsonify({"error": "未登录"}), 401
            return redirect(url_for("login", next=request.path))
        return f(*args, **kwargs)
    return wrapper


# ── CSRF protection ──

def _generate_csrf_token():
    """Generate or retrieve a per-session CSRF token."""
    if "_csrf_token" not in session:
        session["_csrf_token"] = secrets.token_hex(32)
    return session["_csrf_token"]


def _validate_csrf_token(token):
    """Validate the submitted CSRF token against the session token."""
    expected = session.get("_csrf_token", "")
    return hmac.compare_digest(str(token or ""), str(expected))


@app.before_request
def _check_csrf_for_unsafe_methods():
    """CSRF check on POST/PUT/DELETE for routes that require it."""
    if request.method in ("POST", "PUT", "DELETE"):
        # Only enforce CSRF on admin health action routes
        if request.path.startswith("/admin/health/action"):
            token = request.form.get("csrf_token", "")
            if not _validate_csrf_token(token):
                return jsonify({"error": "CSRF token missing or invalid"}), 403


@app.context_processor
def inject_current_user():
    """向所有模板注入当前用户信息。"""
    return {"current_user": session.get("user")}


@app.template_global("csrf_token")
def csrf_token():
    """Template function to expose CSRF token."""
    return _generate_csrf_token()


@app.template_global("group_dates_by_month")
def _group_dates_by_month(all_dates, current_date=None):
    """将日期列表按月份分组，返回月度导航数据。"""
    if not all_dates:
        return None
    months = list(dict.fromkeys(d[:2] for d in all_dates))
    cur_month = current_date[:2] if current_date else (months[-1] if months else '')
    month_dates = [d for d in all_dates if d[:2] == cur_month]
    month_first = {}
    for m in months:
        ds = [d for d in all_dates if d[:2] == m]
        month_first[m] = ds[0]
    idx = months.index(cur_month) if cur_month in months else 0

    def _month_label(mm):
        y = '2026' if mm <= '12' else '2025'
        return f'{y}年{int(mm)}月'

    return {
        'months': months,
        'current_month': cur_month,
        'month_dates': month_dates,
        'month_first': month_first,
        'prev_month': months[idx - 1] if idx > 0 else None,
        'next_month': months[idx + 1] if idx < len(months) - 1 else None,
        'month_label': _month_label(cur_month),
        'month_options': {m: _month_label(m) for m in months},
    }


# ── 登录 / 退出 ──

@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        username = (request.form.get("username") or "").strip()
        password = request.form.get("password") or ""
        ua_info = _parse_ua(request.headers.get("User-Agent", ""))
        client_info = {
            "screen": (request.form.get("_screen") or ""),
            "timezone": (request.form.get("_tz") or ""),
            "platform": (request.form.get("_platform") or ""),
            "language": (request.form.get("_lang") or ""),
            "cores": (request.form.get("_cores") or ""),
            "memory": (request.form.get("_memory") or ""),
            "brands": (request.form.get("_brands") or ""),
            "mobile": (request.form.get("_mobile") or ""),
        }
        user = _find_user(username)
        if not user and '@' in username:
            user = _find_user_by_email(username)
        if not user and '@' not in username:
            user = _find_user_by_email_prefix(username)
        if user and _verify_password(password, user["password_hash"]):
            real_ip = (request.headers.get("X-Real-IP")
                       or (request.headers.get("X-Forwarded-For") or "").split(",")[0].strip()
                       or request.remote_addr
                       or "unknown")
            _log_login(username, user["name"], real_ip, ua_info, client_info)
            session["user"] = {"username": username, "role": user["role"], "name": user["name"]}
            next_page = request.args.get("next") or "/"
            return redirect(next_page)
        return render_template("login.html", error="用户名或密码错误")

    if "user" in session:
        return redirect("/")
    registered = request.args.get("registered")
    return render_template("login.html", success="注册成功，请登录" if registered else "")


@app.route("/logout")
def logout():
    session.pop("user", None)
    return redirect("/login")


# ── 注册 ──

@app.route("/register", methods=["GET", "POST"])
def register():
    if request.method == "POST":
        step = request.form.get("_step", "")
        email = (request.form.get("email") or "").strip().lower()
        code = (request.form.get("code") or "").strip()
        password = request.form.get("password") or ""
        invite = (request.form.get("invite_code") or "").strip()

        # Step 1: 验证邀请码
        if step == "invite":
            if invite != _INVITE_CODE:
                return render_template("register.html", step="invite", error="邀请码错误")
            return render_template("register.html", step="send_code", email=email)

        # Step 2: 发送验证码
        if step == "send_code":
            if not re.match(r"^[^@]+@[^@]+\.[^@]+$", email):
                return render_template("register.html", step="send_code", email=email, error="邮箱格式不正确")
            if _find_user_by_email(email):
                return render_template("register.html", step="send_code", email=email, error="该邮箱已被注册")
            no_resend = _can_resend(email)
            if no_resend:
                return render_template("register.html", step="send_code", email=email, error=no_resend)
            vcode = _gen_code()
            ok, err = _send_code(email, vcode)
            if not ok:
                return render_template("register.html", step="send_code", email=email, error=err)
            _set_code(email, vcode)
            return render_template("register.html", step="verify", email=email,
                                   hint=f"验证码已发送至 {email}，5 分钟内有效")

        # Step 3: 验证码校验 + 创建账户
        if step == "verify":
            if len(password) < 6:
                return render_template("register.html", step="verify", email=email, error="密码长度至少 6 位")
            err = _check_code(email, code)
            if err:
                return render_template("register.html", step="verify", email=email, error=err)
            # 用户名 = 邮箱前缀
            username = email.split("@")[0]
            # 检查用户名冲突
            base_username = username
            suffix = 1
            while _find_user(username) is not None:
                username = f"{base_username}{suffix}"
                suffix += 1
            ok, err = _create_user(username, email, password)
            if not ok:
                return render_template("register.html", step="verify", email=email, error=err)
            return render_template("register.html", step="done", username=username)

        return render_template("register.html", step="invite", error="无效的请求步骤")

    return render_template("register.html", step="invite")


def admin_required(f):
    """仅超级管理员可访问（已包含登录检查）。"""
    @functools.wraps(f)
    def wrapper(*args, **kwargs):
        if "user" not in session:
            if request.is_json or request.path.startswith("/api/"):
                return jsonify({"error": "未登录"}), 401
            return redirect(url_for("login", next=request.path))
        user = session.get("user")
        if not user or user.get("role") != "admin":
            return render_template("login.html", error="无权限，仅管理员可访问"), 403
        return f(*args, **kwargs)
    return wrapper


@app.route("/admin/logins")
@admin_required
def admin_logins():
    records = []
    if os.path.exists(_LOGIN_LOG_PATH):
        try:
            with open(_LOGIN_LOG_PATH, "r", encoding="utf-8") as f:
                records = json.load(f)
        except Exception:
            pass
    return render_template("admin_logins.html", records=list(reversed(records)))


@app.route("/admin/users", methods=["GET", "POST"])
@admin_required
def admin_users():
    msg = ""
    error = ""

    if request.method == "POST":
        action = request.form.get("_action", "")
        target = request.form.get("username", "").strip()
        current = session.get("user", {}).get("username", "")

        if action == "add_user":
            new_username = (request.form.get("new_username") or "").strip()
            new_email = (request.form.get("new_email") or "").strip()
            new_password = (request.form.get("new_password") or "").strip()
            new_role = (request.form.get("new_role") or "user").strip()
            new_name = (request.form.get("new_name") or "").strip()
            if not new_username or not new_email or not new_password:
                error = "用户名、邮箱和密码不能为空"
            elif len(new_password) < 6:
                error = "密码长度至少 6 位"
            else:
                ok, err = _create_user(new_username, new_email, new_password,
                                       role=new_role, name=new_name or new_username)
                if ok:
                    msg = f"已添加用户 {new_username}"
                else:
                    error = err
        elif target == current:
            error = "不能操作自己的账户"
        elif action == "delete":
            if _delete_user(target):
                msg = f"已删除用户 {target}"
            else:
                error = f"删除失败：用户 {target} 不存在"
        elif action == "toggle_role":
            user = _find_user(target)
            if user:
                new_role = "user" if user["role"] == "admin" else "admin"
                _update_user_role(target, new_role)
                msg = f"{target} 角色已切换为 {new_role}"
            else:
                error = f"用户 {target} 不存在"
        elif action == "reset_password":
            new_pw = (request.form.get("new_password") or "").strip()
            if len(new_pw) < 6:
                error = "密码长度至少 6 位"
            elif _reset_user_password(target, new_pw):
                msg = f"已重置 {target} 的密码"
            else:
                error = f"重置失败：用户 {target} 不存在"

    users = _list_users()
    return render_template("admin_users.html", users=users, msg=msg, error=error)


@app.route("/admin/upload", methods=["GET", "POST"])
@admin_required
def admin_upload():
    if request.method == "POST":
        file = request.files.get("file")
        date_str = (request.form.get("date") or "").strip()

        if not re.match(r'^\d{8}$', date_str):
            return jsonify({"error": "日期格式错误，应为 YYYYMMDD"}), 400

        if not file or not file.filename.endswith('.xlsx'):
            return jsonify({"error": "请上传 .xlsx 文件"}), 400

        try:
            df = pd.read_excel(file)
        except Exception as e:
            return jsonify({"error": f"无法读取文件: {e}"}), 400

        if df.empty:
            return jsonify({"error": "文件为空"}), 400

        if len(df) <= 4000:
            return jsonify({"error": f"数据行数不足，当前 {len(df)} 行，需要大于 4000 行"}), 400

        sheet_name = date_str[4:]  # "20260511" -> "0511"

        status = _read_pipeline_status()
        if status.get("running"):
            return jsonify({"error": "流水线正在运行中，请等待完成后再上传"}), 409

        master_path = os.path.join(_PROJECT_ROOT, "Whole Market.xlsx")

        # 保存为临时 CSV，让后台流水线线程写入 Master（避免同步等待）
        temp_csv = os.path.join(_PROJECT_ROOT, f".upload_temp_{date_str}.csv")
        df.to_csv(temp_csv, index=False)

        thread = threading.Thread(
            target=_run_pipeline_for_date,
            args=(date_str, temp_csv),
            daemon=True,
        )
        thread.start()

        return jsonify({"ok": True, "message": f"数据已接收（{len(df)} 行），流水线开始运行"})

    # GET: render upload page
    status = _read_pipeline_status()
    today = datetime.now().strftime("%Y%m%d")

    # Load Guardian health status
    health_status = {}
    health_status_path = os.path.join(os.path.dirname(__file__), "static", "health_status.json")
    last_seen_path = os.path.join(os.path.dirname(__file__), "static", "guardian_last_seen")
    if os.path.exists(health_status_path):
        with open(health_status_path) as f:
            health_status = json.load(f)
    guardian_seconds = -1
    if os.path.exists(last_seen_path):
        with open(last_seen_path) as f:
            guardian_seconds = int(time.time() - float(f.read().strip()))

    return render_template("admin_upload.html",
                           status=status,
                           history=status.get("history", []),
                           today=today,
                           health=health_status,
                           guardian_seconds=guardian_seconds)


@app.route("/api/pipeline/status")
@login_required
def api_pipeline_status():
    status = _read_pipeline_status()
    return jsonify(status)


@app.route("/")
@login_required
def index():
    all_dates = ds.get_available_dates()
    date = request.args.get("date")

    if date and date not in all_dates:
        date = all_dates[-1] if all_dates else None
    if not date and all_dates:
        date = all_dates[-1]

    tracking = ds.get_tracking_data(date) if date else ds.get_latest_tracking()

    prev_date, next_date = ds.get_adjacent_dates(date, all_dates)

    return render_template("index.html",
                           tracking=tracking,
                           all_dates=all_dates,
                           current_date=date,
                           prev_date=prev_date,
                           next_date=next_date)


@app.route("/backtest")
@app.route("/backtest/<date>")
@login_required
def backtest(date=None):
    all_dates = ds.get_available_dates()
    if date and date not in all_dates:
        date = all_dates[-1] if all_dates else None
    if not date and all_dates:
        date = all_dates[-1]

    sheets = ds.get_backtest_sheets(date)

    prev_date, next_date = ds.get_adjacent_dates(date, all_dates)

    return render_template("backtest.html",
                           sheets=sheets,
                           all_dates=all_dates,
                           current_date=date,
                           prev_date=prev_date,
                           next_date=next_date)


@app.route("/weekly")
@app.route("/weekly/<date>")
@login_required
def weekly(date=None):
    all_dates = ds.get_available_weekly_dates()
    if date and date not in all_dates:
        date = all_dates[-1] if all_dates else None
    if not date and all_dates:
        date = all_dates[-1]

    report = ds.get_weekly_data(date)
    sheets, chart_data = ds.extract_weekly_view(report, 'standard')

    weekly_trend = ds.get_weekly_trend_data()

    prev_date, next_date = ds.get_adjacent_dates(date, all_dates)

    return render_template("weekly.html",
                           report=report,
                           sheets=sheets,
                           chart_data=chart_data,
                           weekly_trend=weekly_trend,
                           all_dates=all_dates,
                           current_date=date,
                           prev_date=prev_date,
                           next_date=next_date)


@app.route("/history")
@login_required
def history():
    return render_template("history.html",
                           records=ds.get_history_data(),
                           weekly_strategy=ds.get_weekly_strategy_history(),
                           all_dates=ds.get_available_dates())


@app.route("/screener")
@app.route("/screener/<date>")
@login_required
def screener(date=None):
    all_dates = ds.get_available_dates()
    if date and date not in all_dates:
        date = all_dates[-1] if all_dates else None
    if not date and all_dates:
        date = all_dates[-1]

    industries = ds.get_all_industries()
    total_count = ds.get_screener_count(date)

    prev_date, next_date = ds.get_adjacent_dates(date, all_dates)

    return render_template("screener.html",
                           industries=industries,
                           total_count=total_count,
                           all_dates=all_dates,
                           current_date=date,
                           prev_date=prev_date,
                           next_date=next_date)


# ── 筛选器 AJAX API ──

@app.route("/api/screener/<date>")
@login_required
def api_screener(date):
    """返回分页+筛选后的股票数据 JSON。"""
    acc_min = request.args.get("acc_min", type=float, default=0)
    ind_min = request.args.get("ind_min", type=float, default=0)
    industry = request.args.get("industry", type=str, default="")
    page = request.args.get("page", type=int, default=1)
    per_page = request.args.get("per_page", type=int, default=100)
    per_page = min(per_page, 500)

    result = ds.query_screener(date, acc_min=acc_min, ind_min=ind_min,
                               industry=industry, page=page, per_page=per_page)
    return jsonify(result)


# ── JSON API ──

@app.route("/api/history")
@login_required
def api_history():
    return jsonify(ds.get_history_data())


@app.route("/api/history/weekly-strategy")
@login_required
def api_history_weekly_strategy():
    return jsonify(ds.get_weekly_strategy_history())


@app.route("/api/latest")
@login_required
def api_latest():
    t = ds.get_latest_tracking()
    return jsonify(t) if t else jsonify({"error": "暂无数据"})


@app.route("/api/dates")
@login_required
def api_dates():
    return jsonify(ds.get_available_dates())


@app.route("/api/cache/status")
@login_required
def api_cache_status():
    return jsonify(dbs.get_cache_status())


@app.route("/debug")
def debug():
    """调试页 — 返回数据状态，无任何模板/CSS/JS，纯文本。"""
    dates = ds.get_available_dates()
    track = ds.get_latest_tracking()
    lines = [
        f"Available dates ({len(dates)}): {', '.join(dates)}",
        f"Latest tracking: {json.dumps(track, ensure_ascii=False) if track else 'NONE'}",
    ]
    from datetime import datetime
    lines.append(f"Server time: {datetime.now().isoformat()}")
    return Response("\n".join(lines), mimetype="text/plain")


@app.route("/health")
def health():
    return jsonify({"status": "ok"})


# ── Excel 导出 ──

@app.route("/export/<table>")
@app.route("/export/<table>/<date>")
@login_required
def export_excel(table, date=None):
    """导出表格数据为 Excel (.xlsx)。table: backtest, screener, history, weekly"""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    rows = []
    filename = f"{table}.xlsx"
    sheet_title = "Sheet1"

    if table == "backtest":
        sheets = ds.get_backtest_sheets(date)
        data = sheets.get('backtest')
        if data and data.get('rows'):
            rows = data['rows']
        filename = f"回测明细_{date or 'latest'}.xlsx"
        sheet_title = "回测明细"

    elif table == "screener":
        stocks = ds.get_screener_data(date)
        if stocks:
            rows = stocks
        filename = f"全市场股票_{date or 'latest'}.xlsx"
        sheet_title = "全市场股票"

    elif table == "history":
        records = ds.get_history_data()
        if records:
            rows = records
        filename = "历史趋势.xlsx"
        sheet_title = "历史趋势"

    elif table == "weekly":
        report = ds.get_weekly_data(date)
        sheets, chart_data = ds.extract_weekly_view(report, 'standard')
        if chart_data:
            rows = chart_data
        filename = f"礼拜攻势_{date or 'latest'}.xlsx"
        sheet_title = "礼拜攻势总结"

    if not rows:
        return jsonify({"error": "无数据"}), 404

    cols = list(rows[0].keys())

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_title

    # Header
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="0d6efd", end_color="0d6efd", fill_type="solid")
    for ci, col_name in enumerate(cols, 1):
        cell = ws.cell(row=1, column=ci, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    # Data rows
    for ri, r in enumerate(rows, 2):
        for ci, c in enumerate(cols, 1):
            ws.cell(row=ri, column=ci, value=r.get(c, ''))

    # Auto-adjust column widths
    for ci, col_name in enumerate(cols, 1):
        max_width = len(str(col_name)) * 2
        for ri in range(2, len(rows) + 2):
            val = ws.cell(row=ri, column=ci).value
            if val:
                max_width = max(max_width, len(str(val)) * 1.2)
        ws.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = min(max_width + 4, 50)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    from urllib.parse import quote
    return Response(
        output.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename*=UTF-8''{quote(filename)}"
        },
    )


# ── Admin Health Dashboard ──

@app.route("/admin/health")
@admin_required
def admin_health():
    status_path = os.path.join(os.path.dirname(__file__), "static", "health_status.json")
    last_seen_path = os.path.join(os.path.dirname(__file__), "static", "guardian_last_seen")

    status = {}
    if os.path.exists(status_path):
        with open(status_path) as f:
            status = json.load(f)

    # Guardian freshness check
    guardian_seconds = -1
    guardian_alive = False
    if os.path.exists(last_seen_path):
        with open(last_seen_path) as f:
            last_ts = float(f.read().strip())
        guardian_seconds = int(time.time() - last_ts)
        guardian_alive = guardian_seconds < 900  # 15 minutes

    # Last pipeline run — 日志是中文
    pipeline_log = "/var/log/quant/pipeline.log"
    pipeline_status = {"last_run": "未知", "status": "unknown"}
    if os.path.exists(pipeline_log):
        with open(pipeline_log) as f:
            lines = f.readlines()
        for line in reversed(lines):
            if "流水线" in line and ("完成" in line or "异常" in line):
                pipeline_status["last_run"] = line.strip()
                pipeline_status["status"] = "ok" if "完成" in line else "fail"
                break

    # 用 HTML 模板展示
    return render_template("admin_health.html",
        guardian={"alive": guardian_alive, "seconds_ago": guardian_seconds},
        pipeline=pipeline_status,
        services=status.get("services", {}),
        disk=status.get("disk", {}),
        data_validation=status.get("data_validation", {}),
        docker_uptime=status.get("docker_uptime", {}),
        pipeline_history=status.get("pipeline_history", []),
        backups=status.get("backups", []),
        data_stats=status.get("data_stats", {}),
    )


@app.route("/admin/health/action/<action>", methods=["POST"])
@admin_required
def admin_health_action(action):
    actions = {
        "restart_crawler": ["bash", "/root/sop/repair_crawler.sh"],
        "restart_docker": ["bash", "/root/sop/repair_docker.sh"],
        "backup_now": ["bash", "/root/sop/backup.sh", "daily"],
        "rebuild_sqlite": ["python3", "/var/www/quant/quant_web/refresh_market_cache.py",
                           "--latest"],
    }
    if (action == "rebuild_sqlite_all"
            and market_store.get_storage_mode() is market_store.StorageMode.CSV):
        actions[action] = ["python3", "/var/www/quant/quant_web/import_market_csv.py",
                           "--all"]
    if action == "export_market_xlsx":
        subprocess.Popen(
            [
                sys.executable,
                os.path.join(_PROJECT_ROOT, "quant_web", "export_market_xlsx.py"),
                "--output",
                os.path.join(_PROJECT_ROOT, "Whole Market.xlsx"),
            ],
            start_new_session=True,
            stdout=subprocess.DEVNULL,
            stderr=subprocess.DEVNULL,
        )
        flash("Workbook generation started")
        return redirect(url_for("admin_health"))
    if action in actions:
        subprocess.Popen(actions[action])
        flash("Operation triggered")
    return redirect(url_for("admin_health"))


if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5000, debug=True)
