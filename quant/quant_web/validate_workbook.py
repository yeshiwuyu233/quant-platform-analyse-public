"""
Whole Market.xlsx 完整性校验脚本
用法:  python quant_web/validate_workbook.py
       python quant_web/validate_workbook.py --auto-repair

可重复执行，无副作用。
"""
import argparse
import os
import sys
import glob

PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
MASTER_FILE = os.path.join(PROJECT_ROOT, "Whole Market.xlsx")


def validate(path: str) -> dict:
    """校验 xlsx 可读性，返回统计信息。"""
    import openpyxl
    result = {"ok": False, "sheets": 0, "sheet_names": [], "errors": []}
    try:
        wb = openpyxl.load_workbook(path, read_only=True)
        result["sheets"] = len(wb.sheetnames)
        result["sheet_names"] = wb.sheetnames
        result["size_mb"] = round(os.path.getsize(path) / 1024 / 1024, 1)

        # 检查最后一个 sheet 是否能正常读取
        if wb.sheetnames:
            ws = wb[wb.sheetnames[-1]]
            result["last_sheet_rows"] = ws.max_row
            result["last_sheet_name"] = wb.sheetnames[-1]

        wb.close()
        result["ok"] = True
    except Exception as e:
        result["errors"].append(str(e))

    # 检查是否有 .tmp 残留
    tmp_files = glob.glob(os.path.join(PROJECT_ROOT, "Whole Market*.tmp"))
    result["stale_tmp_files"] = tmp_files if tmp_files else []

    return result


def auto_repair(path: str) -> bool:
    """尝试用 openpyxl 重建 workbook（丢弃损坏部分）。"""
    import openpyxl
    print(f"[自动修复] 正在尝试读取 {path} ...")
    try:
        wb = openpyxl.load_workbook(path)
        repaired = path + ".repaired"
        wb.save(repaired)
        print(f"[自动修复] ✅ 已保存修复版到 {repaired}")
        print(f"[自动修复] 请手动验证后替换原文件：")
        print(f"    mv {repaired} {path}")
        return True
    except Exception as e:
        print(f"[自动修复] ❌ 修复失败: {e}")
        return False


def main():
    parser = argparse.ArgumentParser(description="校验 Whole Market.xlsx")
    parser.add_argument("--auto-repair", action="store_true", help="尝试自动修复损坏文件")
    parser.add_argument("--path", default=MASTER_FILE, help="目标文件路径")
    args = parser.parse_args()

    if not os.path.exists(args.path):
        print(f"❌ 文件不存在: {args.path}")
        sys.exit(1)

    result = validate(args.path)

    print(f"\n{'='*50}")
    print(f"  Whole Market.xlsx 校验报告")
    print(f"{'='*50}")
    print(f"  文件大小: {result.get('size_mb', '?')} MB")
    print(f"  可读性:   {'✅ 正常' if result['ok'] else '❌ 损坏'}")
    print(f"  Sheet数:  {result['sheets']}")
    print(f"  最新Sheet: {result.get('last_sheet_name', '-')} ({result.get('last_sheet_rows', '-')}行)")
    stale = result.get("stale_tmp_files", [])
    print(f"  残留.tmp: {'⚠️ ' + ', '.join(stale) if stale else '✅ 无'}")
    print(f"{'='*50}\n")

    if not result["ok"]:
        for err in result["errors"]:
            print(f"  错误: {err}")
        if args.auto_repair:
            auto_repair(args.path)
        sys.exit(1)

    sys.exit(0)


if __name__ == "__main__":
    main()
